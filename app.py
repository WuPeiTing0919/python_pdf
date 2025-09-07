from fastapi import FastAPI, Request
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import uuid, os

app = FastAPI()

os.makedirs("static", exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")


class TextRequest(BaseModel):
    content: str
    separator: str = "\t"  # 預設使用 Tab 分隔符
    has_header: bool = False  # 是否第一行是標題

@app.post("/generate-excel")
def generate_excel(data: TextRequest, request: Request):
    filename = f"{uuid.uuid4()}.xlsx"
    filepath = os.path.join("static", filename)

    # 創建新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "文字內容"

    # 將文字內容按行分割
    lines = data.content.splitlines()
    
    if not lines:
        return {"error": "沒有提供文字內容"}

    # 分析第一行來確定欄位數量
    first_line = lines[0]
    columns = first_line.split(data.separator)
    num_columns = len(columns)
    
    # 設置標題（如果有指定）
    start_row = 1
    if data.has_header and len(lines) > 1:
        # 第一行作為標題
        for col_idx, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header.strip())
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center')
        start_row = 2
        lines = lines[1:]  # 跳過標題行
    else:
        # 沒有標題，第一行也是資料
        lines = lines

    # 寫入資料
    for row_idx, line in enumerate(lines, start=start_row):
        columns_data = line.split(data.separator)
        
        # 確保每行都有足夠的欄位
        while len(columns_data) < num_columns:
            columns_data.append("")
        
        # 寫入每個欄位的資料
        for col_idx, cell_data in enumerate(columns_data[:num_columns], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_data.strip())
            cell.alignment = Alignment(wrap_text=True)

    # 自動調整列寬
    for col_idx in range(1, num_columns + 1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        max_length = 0
        
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
        # 設置列寬，最小10，最大50
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存文件
    wb.save(filepath)

    return {
        "download_url": str(request.base_url).rstrip("/") + f"/static/{filename}",
        "columns": num_columns,
        "rows": len(lines) + (1 if data.has_header else 0)
    }
